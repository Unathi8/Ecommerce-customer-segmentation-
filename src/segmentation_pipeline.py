"""End-to-end RFM and K-Means customer segmentation pipeline.

Usage:
    python src/segmentation_pipeline.py --input "data/raw/Online Retail.xlsx"

The script accepts an Online Retail-style transactional workbook with these columns:
InvoiceNo, StockCode, Description, Quantity, InvoiceDate, UnitPrice,
CustomerID, and Country.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.preprocessing import StandardScaler

RANDOM_STATE = 42
CLUSTER_COUNT = 4
EXPECTED_COLUMNS = {
    "InvoiceNo",
    "StockCode",
    "Description",
    "Quantity",
    "InvoiceDate",
    "UnitPrice",
    "CustomerID",
    "Country",
}
SEGMENT_ORDER = ["Champions", "Loyal Customers", "Potential Loyalists", "At Risk"]
SEGMENT_COLORS = {
    "Champions": "#0F766E",
    "Loyal Customers": "#2563EB",
    "Potential Loyalists": "#D97706",
    "At Risk": "#DC2626",
}


@dataclass(frozen=True)
class OutputPaths:
    """Project-relative locations for generated analysis artefacts."""

    processed: Path
    figures: Path
    excel: Path


def normalize_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace from headers and harmonize common alternate names."""
    frame = frame.copy()
    frame.columns = frame.columns.astype(str).str.strip()
    aliases = {
        "Invoice No": "InvoiceNo",
        "Invoice Number": "InvoiceNo",
        "Invoice Date": "InvoiceDate",
        "Unit Price": "UnitPrice",
        "Customer ID": "CustomerID",
    }
    return frame.rename(columns=aliases)


def load_and_clean_transactions(input_path: Path) -> tuple[pd.DataFrame, dict[str, int]]:
    """Read and clean invoice-level records for customer segmentation.

    Cancellations and transactions with missing customer identifiers or
    non-positive purchase values are removed because they cannot represent a
    completed customer purchase in the RFM model.
    """
    transactions = normalize_columns(pd.read_excel(input_path))
    missing = EXPECTED_COLUMNS.difference(transactions.columns)
    if missing:
        raise ValueError(
            "Input data is missing required columns: " + ", ".join(sorted(missing))
        )

    audit: dict[str, int] = {"raw_rows": len(transactions)}
    transactions = transactions.loc[:, sorted(EXPECTED_COLUMNS)].copy()
    transactions["InvoiceDate"] = pd.to_datetime(transactions["InvoiceDate"], errors="coerce")
    transactions["CustomerID"] = pd.to_numeric(transactions["CustomerID"], errors="coerce")
    transactions["Quantity"] = pd.to_numeric(transactions["Quantity"], errors="coerce")
    transactions["UnitPrice"] = pd.to_numeric(transactions["UnitPrice"], errors="coerce")
    transactions["InvoiceNo"] = transactions["InvoiceNo"].astype(str).str.strip()

    before = len(transactions)
    transactions = transactions.dropna(
        subset=["InvoiceNo", "InvoiceDate", "CustomerID", "Quantity", "UnitPrice"]
    )
    audit["dropped_missing_critical_fields"] = before - len(transactions)

    before = len(transactions)
    cancellation_mask = transactions["InvoiceNo"].str.upper().str.startswith("C")
    transactions = transactions.loc[~cancellation_mask].copy()
    audit["dropped_cancellations"] = before - len(transactions)

    before = len(transactions)
    transactions = transactions.loc[
        (transactions["Quantity"] > 0) & (transactions["UnitPrice"] > 0)
    ].copy()
    audit["dropped_non_positive_transactions"] = before - len(transactions)

    before = len(transactions)
    transactions = transactions.drop_duplicates()
    audit["dropped_duplicate_lines"] = before - len(transactions)

    transactions["CustomerID"] = transactions["CustomerID"].astype("int64").astype(str)
    transactions["Revenue"] = (transactions["Quantity"] * transactions["UnitPrice"]).round(2)
    transactions = transactions.sort_values(["CustomerID", "InvoiceDate", "InvoiceNo"]).reset_index(drop=True)
    audit["clean_rows"] = len(transactions)
    audit["customers_after_cleaning"] = transactions["CustomerID"].nunique()
    return transactions, audit


def most_frequent_or_unknown(values: pd.Series) -> str:
    """Return the modal country, handling empty values safely."""
    modes = values.dropna().mode()
    return str(modes.iloc[0]) if not modes.empty else "Unknown"


def build_customer_features(transactions: pd.DataFrame) -> tuple[pd.DataFrame, pd.Timestamp]:
    """Aggregate cleaned transaction lines into customer-level RFM features."""
    snapshot_date = pd.Timestamp(transactions["InvoiceDate"].max()).normalize() + pd.offsets.Day(1)
    customers = (
        transactions.groupby("CustomerID", as_index=False)
        .agg(
            LastPurchaseDate=("InvoiceDate", "max"),
            FirstPurchaseDate=("InvoiceDate", "min"),
            Frequency=("InvoiceNo", "nunique"),
            Monetary=("Revenue", "sum"),
            UnitsPurchased=("Quantity", "sum"),
            Country=("Country", most_frequent_or_unknown),
        )
        .copy()
    )
    customers["Recency"] = (snapshot_date - customers["LastPurchaseDate"].dt.normalize()).dt.days
    customers["TenureDays"] = (
        customers["LastPurchaseDate"].dt.normalize() - customers["FirstPurchaseDate"].dt.normalize()
    ).dt.days
    customers["AverageOrderValue"] = customers["Monetary"] / customers["Frequency"]
    customers["Monetary"] = customers["Monetary"].round(2)
    customers["AverageOrderValue"] = customers["AverageOrderValue"].round(2)
    customers = customers[
        [
            "CustomerID",
            "Country",
            "FirstPurchaseDate",
            "LastPurchaseDate",
            "Recency",
            "Frequency",
            "Monetary",
            "AverageOrderValue",
            "UnitsPurchased",
            "TenureDays",
        ]
    ].sort_values("CustomerID").reset_index(drop=True)
    return customers, snapshot_date


def prepare_rfm_matrix(customers: pd.DataFrame) -> tuple[np.ndarray, StandardScaler]:
    """Log-transform skewed RFM values and standardize them for K-Means."""
    rfm_values = customers[["Recency", "Frequency", "Monetary"]].astype(float)
    transformed = np.log1p(rfm_values)
    scaler = StandardScaler()
    return scaler.fit_transform(transformed), scaler


def calculate_model_selection_scores(
    scaled_rfm: np.ndarray, k_values: Iterable[int] = range(2, 9)
) -> pd.DataFrame:
    """Calculate elbow inertia and silhouette diagnostics across plausible k values."""
    results: list[dict[str, float | int]] = []
    for k in k_values:
        model = KMeans(n_clusters=k, random_state=RANDOM_STATE, n_init=25)
        labels = model.fit_predict(scaled_rfm)
        results.append(
            {
                "Clusters": k,
                "Inertia": round(float(model.inertia_), 2),
                "SilhouetteScore": round(float(silhouette_score(scaled_rfm, labels)), 4),
            }
        )
    return pd.DataFrame(results)


def assign_segment_names(customers: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Map anonymous K-Means labels to interpretable business segments.

    The mapping is deterministic and based on each cluster's unrounded RFM
    profile. The strongest current-value group becomes Champions; the least
    recently active remaining group becomes At Risk; the highest remaining
    frequency/monetary group becomes Loyal Customers; the final group becomes
    Potential Loyalists.
    """
    profiles = (
        customers.groupby("Cluster", as_index=False)
        .agg(
            Customers=("CustomerID", "nunique"),
            AvgRecencyDays=("Recency", "mean"),
            AvgFrequency=("Frequency", "mean"),
            AvgMonetary=("Monetary", "mean"),
            AvgOrderValue=("AverageOrderValue", "mean"),
            AvgTenureDays=("TenureDays", "mean"),
        )
        .copy()
    )
    profiles["RecencyScore"] = 1 - profiles["AvgRecencyDays"].rank(pct=True, method="first")
    profiles["FrequencyScore"] = profiles["AvgFrequency"].rank(pct=True, method="first")
    profiles["MonetaryScore"] = profiles["AvgMonetary"].rank(pct=True, method="first")
    profiles["CurrentValueScore"] = (
        0.40 * profiles["RecencyScore"]
        + 0.30 * profiles["FrequencyScore"]
        + 0.30 * profiles["MonetaryScore"]
    )

    champions_cluster = int(profiles.loc[profiles["CurrentValueScore"].idxmax(), "Cluster"])
    remaining = profiles.loc[profiles["Cluster"] != champions_cluster].copy()
    at_risk_cluster = int(remaining.loc[remaining["AvgRecencyDays"].idxmax(), "Cluster"])
    remaining = remaining.loc[remaining["Cluster"] != at_risk_cluster].copy()
    loyal_cluster = int(
        remaining.loc[
            (remaining["FrequencyScore"] + remaining["MonetaryScore"]).idxmax(), "Cluster"
        ]
    )
    potential_cluster = int(
        remaining.loc[remaining["Cluster"] != loyal_cluster, "Cluster"].iloc[0]
    )

    cluster_to_segment = {
        champions_cluster: "Champions",
        loyal_cluster: "Loyal Customers",
        potential_cluster: "Potential Loyalists",
        at_risk_cluster: "At Risk",
    }
    segmented = customers.copy()
    segmented["Segment"] = segmented["Cluster"].map(cluster_to_segment)
    segmented["Segment"] = pd.Categorical(segmented["Segment"], categories=SEGMENT_ORDER, ordered=True)
    segment_sort = {segment: order for order, segment in enumerate(SEGMENT_ORDER, start=1)}
    segmented["SegmentOrder"] = segmented["Segment"].map(segment_sort).astype(int)

    segment_profiles = (
        segmented.groupby("Segment", observed=False, as_index=False)
        .agg(
            Customers=("CustomerID", "nunique"),
            AvgRecencyDays=("Recency", "mean"),
            AvgFrequency=("Frequency", "mean"),
            AvgMonetary=("Monetary", "mean"),
            TotalRevenue=("Monetary", "sum"),
            AvgOrderValue=("AverageOrderValue", "mean"),
            AvgTenureDays=("TenureDays", "mean"),
        )
        .copy()
    )
    segment_profiles["CustomerSharePct"] = 100 * segment_profiles["Customers"] / segment_profiles["Customers"].sum()
    segment_profiles["RevenueSharePct"] = 100 * segment_profiles["TotalRevenue"] / segment_profiles["TotalRevenue"].sum()
    segment_profiles["SegmentOrder"] = segment_profiles["Segment"].map(segment_sort).astype(int)
    segment_profiles["RecommendedAction"] = segment_profiles["Segment"].map(
        {
            "Champions": "Reward loyalty with VIP access, premium bundles, and referral incentives.",
            "Loyal Customers": "Use cross-sell offers and loyalty milestones to raise customer lifetime value.",
            "Potential Loyalists": "Encourage the next purchase with personalized product discovery and time-bound offers.",
            "At Risk": "Run a targeted win-back campaign using a compelling reactivation offer and feedback request.",
        }
    )
    numeric_columns = segment_profiles.select_dtypes(include="number").columns
    segment_profiles[numeric_columns] = segment_profiles[numeric_columns].round(2)
    return segmented, segment_profiles


def fit_customer_segments(customers: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Fit four K-Means clusters and return assigned customers plus diagnostics."""
    scaled_rfm, _ = prepare_rfm_matrix(customers)
    model_selection = calculate_model_selection_scores(scaled_rfm)
    model = KMeans(n_clusters=CLUSTER_COUNT, random_state=RANDOM_STATE, n_init=25)
    assigned = customers.copy()
    assigned["Cluster"] = model.fit_predict(scaled_rfm)
    assigned["Cluster"] = assigned["Cluster"].astype(int)
    assigned, segment_profiles = assign_segment_names(assigned)
    return assigned, segment_profiles, model_selection


def create_figures(
    customers: pd.DataFrame, segment_profiles: pd.DataFrame, model_selection: pd.DataFrame, figure_dir: Path
) -> None:
    """Create presentation-quality charts for the README and project report."""
    figure_dir.mkdir(parents=True, exist_ok=True)
    sns.set_theme(style="whitegrid", font_scale=1.0)

    fig, ax = plt.subplots(figsize=(8.5, 5.2), dpi=160)
    ax.plot(model_selection["Clusters"], model_selection["Inertia"], marker="o", color="#2563EB", linewidth=2.5)
    ax.axvline(CLUSTER_COUNT, color="#DC2626", linestyle="--", linewidth=1.5, label="Selected k = 4")
    ax.set_title("Elbow Method: K-Means Inertia by Cluster Count", weight="bold", pad=12)
    ax.set_xlabel("Number of clusters (k)")
    ax.set_ylabel("Within-cluster sum of squares (inertia)")
    ax.set_xticks(model_selection["Clusters"])
    ax.legend(frameon=True)
    fig.tight_layout()
    fig.savefig(figure_dir / "elbow_method.png", bbox_inches="tight")
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(9, 6), dpi=160)
    for segment in SEGMENT_ORDER:
        subset = customers.loc[customers["Segment"] == segment]
        ax.scatter(
            subset["Recency"],
            subset["Monetary"],
            label=segment,
            color=SEGMENT_COLORS[segment],
            alpha=0.58,
            s=22,
            edgecolors="none",
        )
    ax.set_yscale("log")
    ax.set_title("Customer Segments by Recency and Monetary Value", weight="bold", pad=12)
    ax.set_xlabel("Recency (days since last purchase)")
    ax.set_ylabel("Monetary value (GBP, logarithmic scale)")
    ax.legend(title="Segment", frameon=True)
    fig.tight_layout()
    fig.savefig(figure_dir / "segment_scatter.png", bbox_inches="tight")
    plt.close(fig)

    heatmap_data = segment_profiles.set_index("Segment")[
        ["AvgRecencyDays", "AvgFrequency", "AvgMonetary", "AvgOrderValue"]
    ].copy()
    heatmap_scaled = (heatmap_data - heatmap_data.mean()) / heatmap_data.std(ddof=0)
    fig, ax = plt.subplots(figsize=(9, 4.6), dpi=160)
    sns.heatmap(
        heatmap_scaled,
        cmap="RdYlGn",
        center=0,
        annot=heatmap_data.round(1),
        fmt=",.1f",
        linewidths=0.8,
        cbar_kws={"label": "Relative value within metric"},
        ax=ax,
    )
    ax.set_title("Segment Profile Comparison", weight="bold", pad=12)
    ax.set_xlabel("Metric")
    ax.set_ylabel("Segment")
    fig.tight_layout()
    fig.savefig(figure_dir / "segment_profile_heatmap.png", bbox_inches="tight")
    plt.close(fig)


def create_recommendations() -> pd.DataFrame:
    """Return concise, dashboard-friendly campaign recommendations."""
    return pd.DataFrame(
        [
            {
                "Segment": "Champions",
                "BusinessObjective": "Protect and grow highest-value relationships",
                "RecommendedTactic": "VIP early access, referral rewards, premium bundles",
                "PrimaryKPI": "Repeat revenue and referral conversion",
            },
            {
                "Segment": "Loyal Customers",
                "BusinessObjective": "Increase basket size and purchase cadence",
                "RecommendedTactic": "Cross-sell recommendations and loyalty milestones",
                "PrimaryKPI": "Average order value and purchase frequency",
            },
            {
                "Segment": "Potential Loyalists",
                "BusinessObjective": "Convert recent buyers into repeat customers",
                "RecommendedTactic": "Personalized discovery email with time-bound incentive",
                "PrimaryKPI": "Second-purchase conversion",
            },
            {
                "Segment": "At Risk",
                "BusinessObjective": "Reactivate inactive formerly valuable customers",
                "RecommendedTactic": "Win-back sequence with offer and feedback request",
                "PrimaryKPI": "Reactivation rate and recovered revenue",
            },
        ]
    )


def save_excel_report(
    customers: pd.DataFrame,
    segment_profiles: pd.DataFrame,
    model_selection: pd.DataFrame,
    recommendations: pd.DataFrame,
    audit: dict[str, int],
    snapshot_date: pd.Timestamp,
    output_path: Path,
) -> None:
    """Write formatted Excel workbook for stakeholders and Power BI handoff."""
    summary = pd.DataFrame(
        [
            ("Analysis snapshot date", snapshot_date.date().isoformat()),
            ("Customers segmented", int(customers["CustomerID"].nunique())),
            ("Completed transaction lines", int(audit["clean_rows"])),
            ("Total customer revenue", float(customers["Monetary"].sum())),
            ("Selected clusters", CLUSTER_COUNT),
            ("Clustering method", "K-Means on standardized log-transformed RFM metrics"),
            ("Random seed", RANDOM_STATE),
        ],
        columns=["Metric", "Value"],
    )
    audit_frame = pd.DataFrame(audit.items(), columns=["CleaningMetric", "RowsOrCustomers"])
    dictionary = pd.DataFrame(
        [
            ("CustomerID", "Unique customer identifier"),
            ("Recency", "Days between analysis snapshot and last completed purchase"),
            ("Frequency", "Distinct completed invoices for the customer"),
            ("Monetary", "Total completed transaction revenue in GBP"),
            ("AverageOrderValue", "Monetary value divided by distinct completed invoices"),
            ("Segment", "Business-friendly label mapped from the K-Means cluster"),
        ],
        columns=["Field", "Definition"],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        summary.to_excel(writer, sheet_name="Executive Summary", index=False)
        segment_profiles.to_excel(writer, sheet_name="Segment Profiles", index=False)
        recommendations.to_excel(writer, sheet_name="Recommendations", index=False)
        customers.to_excel(writer, sheet_name="Customer Segments", index=False)
        model_selection.to_excel(writer, sheet_name="Model Selection", index=False)
        audit_frame.to_excel(writer, sheet_name="Cleaning Audit", index=False)
        dictionary.to_excel(writer, sheet_name="Data Dictionary", index=False)

    workbook = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 48)
    profile_sheet = workbook["Segment Profiles"]
    profile_sheet.conditional_formatting.add(
        f"C2:G{profile_sheet.max_row}",
        ColorScaleRule(start_type="min", start_color="FEE2E2", mid_type="percentile", mid_value=50, mid_color="FEF3C7", end_type="max", end_color="DCFCE7"),
    )
    workbook.save(output_path)


def export_outputs(
    transactions: pd.DataFrame,
    customers: pd.DataFrame,
    segment_profiles: pd.DataFrame,
    model_selection: pd.DataFrame,
    recommendations: pd.DataFrame,
    audit: dict[str, int],
    snapshot_date: pd.Timestamp,
    output_paths: OutputPaths,
) -> None:
    """Persist CSV, figures, and Excel outputs required by the project."""
    output_paths.processed.mkdir(parents=True, exist_ok=True)
    customers.to_csv(output_paths.processed / "customer_segments.csv", index=False)
    segment_profiles.to_csv(output_paths.processed / "segment_profiles.csv", index=False)
    model_selection.to_csv(output_paths.processed / "model_selection_metrics.csv", index=False)
    recommendations.to_csv(output_paths.processed / "segment_recommendations.csv", index=False)
    pd.DataFrame(audit.items(), columns=["CleaningMetric", "RowsOrCustomers"]).to_csv(
        output_paths.processed / "cleaning_audit.csv", index=False
    )
    create_figures(customers, segment_profiles, model_selection, output_paths.figures)
    save_excel_report(
        customers,
        segment_profiles,
        model_selection,
        recommendations,
        audit,
        snapshot_date,
        output_paths.excel,
    )


def main() -> None:
    """Run the complete segmentation workflow from the command line."""
    parser = argparse.ArgumentParser(description="Run e-commerce RFM customer segmentation.")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/raw/Online Retail.xlsx"),
        help="Path to an Online Retail-style Excel input file.",
    )
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path("."),
        help="Project root where data, reports, and outputs folders will be written.",
    )
    args = parser.parse_args()
    input_path = args.input.expanduser().resolve()
    root = args.project_root.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(
            f"Input workbook was not found at {input_path}. Download the UCI Online Retail data first."
        )

    transactions, audit = load_and_clean_transactions(input_path)
    customer_features, snapshot_date = build_customer_features(transactions)
    customers, segment_profiles, model_selection = fit_customer_segments(customer_features)
    recommendations = create_recommendations()
    export_outputs(
        transactions,
        customers,
        segment_profiles,
        model_selection,
        recommendations,
        audit,
        snapshot_date,
        OutputPaths(
            processed=root / "data" / "processed",
            figures=root / "reports" / "figures",
            excel=root / "outputs" / "customer_segmentation_report.xlsx",
        ),
    )
    print("Customer segmentation completed successfully.")
    print(f"Customers segmented: {customers['CustomerID'].nunique():,}")
    print(f"Total revenue analyzed: GBP {customers['Monetary'].sum():,.2f}")
    print("Selected model: K-Means with k=4 (random state 42)")


if __name__ == "__main__":
    main()
